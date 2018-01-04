# -*- coding: utf-8 -*-
"""
Created on Tue Jan  2 21:42:41 2018

@author: cosmo

Creates an interactive graph that reads in hierarchical data and allows
the user to visualize the components of the data structure
"""

import sys, os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from helper_functions import search_xl
from plot_tree import Plot_tree


def tree_from_excel(path_to_file, sheet_name):
    # construct a plot_tree from excel file
    
    # read in the values from excel file
    wb_path = path_to_file
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb.get_sheet_by_name(sheet_name)
    
    c = search_xl(ws, "Hierarchy")
    hier = []
    cell = ws[get_column_letter(c[1])+str(c[0]+1)]
    hier.append(cell.value)
    cell = ws[cell.column+str(cell.row+1)]
    while not cell.value is None:
        if (cell.value > hier[-1] + 1) or cell.value < 1:
            print("invalid hierarchy!")
            break
        hier.append(cell.value)
        cell = ws[cell.column+str(cell.row+1)]
    print(hier)
    
    c = search_xl(ws, "Index")
    cols = []
    idx = []
    col_num = c[1] + 1
    cell = ws[get_column_letter(col_num)+str(c[0])]
    
    while not cell.value is None:
        cols.append(col_num)
        idx.append(cell.value)
        col_num = col_num + 1
        cell = ws[get_column_letter(col_num)+str(cell.row)]
    print(idx)
        
    df = pd.DataFrame(columns = idx)
    names = []
    cell = ws[get_column_letter(c[1])+str(c[0]+1)]
    
    while not cell.value is None:
        names.append(cell.value)
        
        new_row = []
        for col in cols:
            new_row.append(ws[get_column_letter(col)+str(cell.row)].value)
        df.loc[cell.value] = new_row
        
        cell = ws[cell.column+str(cell.row+1)]
    print(names)
    
    # construct the Plot_tree
    tree_ladder = []
    for i, lvl in enumerate(hier):
        new_name = names[i]
        new_row = df.loc[new_name]
        
        new_tree = Plot_tree(idx, new_row, new_name)
        
        while True:
            if lvl == 0:
                tree_ladder.append(new_tree)
                base_tree = new_tree
                break
            if lvl > len(tree_ladder)-1:
                new_tree.add_to_parent(tree_ladder[-1])
                tree_ladder.append(new_tree)
                break
            elif lvl == len(tree_ladder)-1:
                new_tree.add_to_parent(tree_ladder[lvl-1])
                tree_ladder[-1] = new_tree
                break
            else:
                tree_ladder.pop(-1)
    
    return base_tree

    
def navigate_tree(base_tree):
    # interactively navigate a given tree
    
    current_tree = base_tree
    
    while True:
        print()
        print("choose a path... ")
        for child in current_tree.get_children():
            print(child)
        print("up (go up one level)")
        print("end (exit the program)")
        
        new_path = input(": ")
        
        if new_path == "up":
            if current_tree.get_parent():
                current_tree = current_tree.get_parent()
            else:
                current_tree.plot_self()
        elif new_path == "end":
            break
        else:
            try:
                current_tree = current_tree.children[new_path]
            except:
                print("please choose a valid path")
                continue
        
        if bool(current_tree.children):
            current_tree.plot_children()
        else:
            current_tree.plot_self()


if __name__ == '__main__':
    base_tree = tree_from_excel("Excel_tree_example.xlsx", "Sheet1")
    
    # plot base tree and children
    
    base_tree.plot_self()
    base_tree.plot_children()
    
    navigate_tree(base_tree)

